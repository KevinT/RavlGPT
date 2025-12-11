#!/usr/bin/env python3
"""
Google Sheets Analyzer Workflow

Handles extraction and analysis of data from Google Sheets and Excel files.

Inherits from GoogleAPIsMixin to access Google Sheets service.
"""

import re
import sys
from datetime import datetime
from typing import Dict, Any, List
from io import BytesIO

# Import shared markdown conversion utilities
from google_sheets_markdown_converter import (
    sheets_to_markdown,
    values_to_markdown_table,
    is_empty_row,
    trim_empty_columns
)


class GoogleSheetsAnalyzer:
    """
    Analyzes and extracts data from Google Sheets.

    Usage:
        analyzer = GoogleSheetsAnalyzer(loop)  # loop must inherit GoogleAPIsMixin
        sheet_data = analyzer.fetch(url)
        markdown = analyzer.export_as_markdown(url)
    """

    def __init__(self, loop_with_mixin):
        """
        Initialize with a loop that has GoogleAPIsMixin.

        Args:
            loop_with_mixin: A RAVL loop instance that inherits GoogleAPIsMixin
        """
        self.loop = loop_with_mixin

    def fetch(self, url: str) -> Dict[str, Any]:
        """
        Fetch content from Google Sheets using the API.

        Args:
            url: Google Sheets URL

        Returns:
            Dict with:
            - 'text': Markdown formatted representation of all sheets
            - 'title': Spreadsheet title
            - 'last_modified': ISO timestamp of last modification
            - 'spreadsheet_id': Extracted spreadsheet ID from URL
            - 'sheets': List of sheet data (name, values)

        Raises:
            ValueError: If URL is invalid
            Exception: If API request fails
        """
        if not self.loop.google_sheets_service or not self.loop.google_drive_service:
            self.loop.init_google_services()

        # Extract spreadsheet ID from URL
        # Format: https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit...
        spreadsheet_id_match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
        if not spreadsheet_id_match:
            raise ValueError(f"Invalid Google Sheets URL: {url}")

        spreadsheet_id = spreadsheet_id_match.group(1)

        try:
            # Detect file type using Drive API
            file_metadata = self.loop.google_drive_service.files().get(
                fileId=spreadsheet_id,
                fields='mimeType,name,modifiedTime',
                supportsAllDrives=True
            ).execute()

            mime_type = file_metadata.get('mimeType')

            # Branch based on file type
            if mime_type == 'application/vnd.google-apps.spreadsheet':
                # Native Google Sheets - use Sheets API
                return self._fetch_native_sheets(spreadsheet_id, file_metadata)
            elif mime_type in [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
                'application/vnd.ms-excel'  # .xls
            ]:
                # Excel file - download and parse with openpyxl
                return self._fetch_excel_file(spreadsheet_id, file_metadata)
            else:
                raise ValueError(f"Unsupported spreadsheet MIME type: {mime_type}")

        except Exception as e:
            raise Exception(f"Failed to fetch spreadsheet: {e}")

    def _fetch_native_sheets(self, spreadsheet_id: str, file_metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Fetch native Google Sheets using Sheets API.

        Args:
            spreadsheet_id: The spreadsheet ID
            file_metadata: File metadata from Drive API

        Returns:
            Dict with text, title, last_modified, spreadsheet_id, sheets
        """
        try:
            # Get spreadsheet metadata and all sheets
            spreadsheet = self.loop.google_sheets_service.spreadsheets().get(
                spreadsheetId=spreadsheet_id
            ).execute()

            title = spreadsheet.get('properties', {}).get('title', 'Untitled Spreadsheet')
            sheets = spreadsheet.get('sheets', [])

            # Get modification timestamp from Drive API metadata
            last_modified = file_metadata.get('modifiedTime')

            # Fetch data for each sheet
            sheets_data = []
            for sheet in sheets:
                sheet_properties = sheet.get('properties', {})
                sheet_title = sheet_properties.get('title', 'Untitled Sheet')
                sheet_id = sheet_properties.get('sheetId')

                try:
                    # Get all values for this sheet
                    result = self.loop.google_sheets_service.spreadsheets().values().get(
                        spreadsheetId=spreadsheet_id,
                        range=sheet_title  # Use sheet title as range to get all data
                    ).execute()

                    values = result.get('values', [])

                    sheets_data.append({
                        'title': sheet_title,
                        'sheet_id': sheet_id,
                        'values': values
                    })

                except Exception as e:
                    from pathlib import Path
                    _utils_dir = Path(__file__).parent.parent / 'utils'
                    import sys
                    if str(_utils_dir) not in sys.path:
                        sys.path.insert(0, str(_utils_dir))
                    from logging_utils import log_execution
                    log_execution(f"Could not fetch data for sheet '{sheet_title}': {e}", status='error')
                    sheets_data.append({
                        'title': sheet_title,
                        'sheet_id': sheet_id,
                        'values': [],
                        'error': str(e)
                    })

            # Convert to markdown
            markdown_text = sheets_to_markdown(title, sheets_data)

            return {
                'text': markdown_text,
                'title': title,
                'last_modified': last_modified,
                'spreadsheet_id': spreadsheet_id,
                'sheets': sheets_data
            }

        except Exception as e:
            raise Exception(f"Failed to fetch native Google Sheets: {e}")

    def _fetch_excel_file(self, spreadsheet_id: str, file_metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Fetch Excel file (.xlsx/.xls) from Google Drive and parse with openpyxl.

        Args:
            spreadsheet_id: The file ID in Google Drive
            file_metadata: File metadata from Drive API

        Returns:
            Dict with text, title, last_modified, spreadsheet_id, sheets
        """
        try:
            # Import openpyxl
            try:
                import openpyxl
            except ImportError:
                raise Exception(
                    "openpyxl is required to parse Excel files. "
                    "Add it to allowed_dependencies in config/ravl.toml"
                )

            # Download Excel file as binary
            request = self.loop.google_drive_service.files().get_media(
                fileId=spreadsheet_id,
                supportsAllDrives=True
            )
            excel_bytes = request.execute()

            # Parse with openpyxl
            workbook = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)

            title = file_metadata.get('name', 'Untitled Spreadsheet')
            last_modified = file_metadata.get('modifiedTime')

            # Extract data from all sheets
            sheets_data = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Convert worksheet to 2D array of values, filtering empty rows
                values = []
                for row in worksheet.iter_rows(values_only=True):
                    # Convert None to empty string and all values to strings
                    row_values = [str(cell) if cell is not None else '' for cell in row]

                    # Skip completely empty rows
                    if not is_empty_row(row_values):
                        values.append(row_values)

                # Trim trailing empty columns from all rows
                if values:
                    values = trim_empty_columns(values)

                sheets_data.append({
                    'title': sheet_name,
                    'sheet_id': None,  # Excel doesn't have sheet IDs
                    'values': values
                })

            workbook.close()

            # Convert to markdown
            markdown_text = sheets_to_markdown(title, sheets_data)

            return {
                'text': markdown_text,
                'title': title,
                'last_modified': last_modified,
                'spreadsheet_id': spreadsheet_id,
                'sheets': sheets_data
            }

        except Exception as e:
            raise Exception(f"Failed to fetch Excel file: {e}")


    def export_as_markdown(self, url: str) -> str:
        """
        Fetch sheets and return as markdown string.

        Args:
            url: Google Sheets URL

        Returns:
            Markdown formatted string

        Raises:
            ValueError: If URL is invalid
            Exception: If API request fails
        """
        sheet_data = self.fetch(url)
        return sheet_data['text']

