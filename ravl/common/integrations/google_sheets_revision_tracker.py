#!/usr/bin/env python3
"""
Google Sheets Revision Tracker Workflow

Handles tracking and exporting revision history from Google Sheets.
Follows the same pattern as GoogleDocsRevisionTracker but exports revisions
as Excel files, then converts to markdown tables.

Inherits from GoogleAPIsMixin to access Google Sheets and Drive services.
"""

import os
import re
import sys
import json
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List, Optional
from io import BytesIO

# Add utils to path for logging
import sys
from pathlib import Path
_utils_dir = Path(__file__).parent.parent / 'utils'
if str(_utils_dir) not in sys.path:
    sys.path.insert(0, str(_utils_dir))
from logging_utils import log_execution, log_message

# Import shared markdown conversion utilities
from google_sheets_markdown_converter import (
    sheets_to_markdown,
    values_to_markdown_table,
    is_empty_row,
    trim_empty_columns
)


class GoogleSheetsRevisionTracker:
    """
    Tracks revision history of Google Sheets with configurable export modes.

    Similar to GoogleDocsRevisionTracker but adapted for spreadsheet formats:
    - Exports revisions as Excel (.xlsx) via Drive API exportLinks
    - Converts Excel to markdown tables matching GoogleSheetsAnalyzer format
    - Tracks sequence numbers and Google revision IDs in metadata.jsonl
    - Implements incremental saving (only new revisions)

    Usage:
        tracker = GoogleSheetsRevisionTracker(loop)  # loop must inherit GoogleAPIsMixin
        revisions = tracker.fetch_revisions(spreadsheet_id, max_count=100)
        tracker.save_missing_revisions(
            document_name="my-sheet",
            spreadsheet_id="abc123",
            output_path="/data/output/",
            revisions_metadata=edit_history,
            max_count=100
        )
    """

    def __init__(self, loop_with_mixin):
        """
        Initialize with a loop that has GoogleAPIsMixin.

        Args:
            loop_with_mixin: A RAVL loop instance that inherits GoogleAPIsMixin
        """
        self.loop = loop_with_mixin
        self.drive_service = None

    def _init_drive_service(self):
        """Initialize Google Drive service for revision access."""
        if not self.drive_service:
            # Reuse credentials from existing Google Sheets service
            if not self.loop.google_sheets_service:
                self.loop.init_google_services()

            # Get credentials from the sheets service
            from googleapiclient.discovery import build

            # Build Drive service with same credentials
            self.drive_service = build('drive', 'v3', http=self.loop.google_sheets_service._http)

    def get_edit_history(self, spreadsheet_id: str, max_count: int = 100) -> List[Dict[str, Any]]:
        """
        Fetch edit history metadata without exporting revision content.

        Used for metadata collection. Returns author, timestamp, and change info.

        Args:
            spreadsheet_id: Google Sheets spreadsheet ID
            max_count: Maximum number of revisions to fetch

        Returns:
            List of edit history entries with:
            - timestamp: ISO timestamp of edit
            - author: Display name of editor
            - email: Email address of editor
            - revision_id: Internal Google revision ID

        Raises:
            Exception: If API request fails
        """
        revisions = self.fetch_revisions(spreadsheet_id, max_count)

        # Extract just the metadata we need for audit trail
        edit_history = []
        for rev in revisions:
            edit_history.append({
                'timestamp': rev.get('modifiedTime'),
                'author': rev.get('lastModifyingUser', {}).get('displayName', 'Unknown'),
                'email': rev.get('lastModifyingUser', {}).get('emailAddress', 'unknown@example.com'),
                'revision_id': rev.get('id')
            })

        return edit_history

    def fetch_revisions(self, spreadsheet_id: str, max_count: int = 100) -> List[Dict[str, Any]]:
        """
        Fetch revision history from Google Drive API.

        Args:
            spreadsheet_id: Google Sheets spreadsheet ID
            max_count: Maximum number of revisions to fetch

        Returns:
            List of revision objects with metadata:
            - id: Revision ID
            - mimeType: MIME type
            - modifiedTime: ISO timestamp
            - lastModifyingUser: User info (displayName, emailAddress)
            - size: File size in bytes
            - keepForever: Whether revision is kept indefinitely

        Raises:
            Exception: If API request fails
        """
        if not self.drive_service:
            self._init_drive_service()

        try:
            results = self.drive_service.revisions().list(
                fileId=spreadsheet_id,
                pageSize=max_count,
                fields='revisions(id, mimeType, modifiedTime, lastModifyingUser, size, keepForever)'
            ).execute()

            revisions = results.get('revisions', [])
            log_execution(f"Fetched {len(revisions)} revisions for spreadsheet {spreadsheet_id}", status='success')
            return revisions

        except Exception as e:
            log_execution(f"Error fetching revisions: {e}", status='error')
            raise

    def export_revision(self, spreadsheet_id: str, revision_id: str, max_retries: int = 3) -> str:
        """
        Export a specific Sheets revision as markdown with retry logic.

        Sheets-specific implementation:
        1. Get revision metadata with exportLinks from Drive API
        2. Download Excel export using exportLinks['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']
        3. Parse Excel bytes with openpyxl (same as GoogleSheetsAnalyzer._fetch_excel_file)
        4. Convert all sheets to markdown tables
        5. Return markdown string

        Args:
            spreadsheet_id: Google Sheets spreadsheet ID
            revision_id: Revision ID to export
            max_retries: Maximum number of retry attempts (default: 3)

        Returns:
            Markdown content of the revision with all sheets as tables

        Raises:
            Exception: If API request fails after all retries
        """
        import time

        if not self.drive_service:
            self._init_drive_service()

        last_error = None
        for attempt in range(max_retries):
            try:
                # Import openpyxl for Excel parsing
                try:
                    import openpyxl
                except ImportError:
                    raise Exception(
                        "openpyxl is required to parse Excel files. "
                        "Add it to allowed_dependencies in config/ravl.toml"
                    )

                # Get revision metadata including exportLinks
                revision = self.drive_service.revisions().get(
                    fileId=spreadsheet_id,
                    revisionId=revision_id,
                    fields='exportLinks'
                ).execute()

                # Get Excel export link
                export_links = revision.get('exportLinks', {})
                xlsx_url = export_links.get('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

                if not xlsx_url:
                    # Try CSV as fallback (only exports first sheet but better than nothing)
                    csv_url = export_links.get('text/csv')
                    if csv_url:
                        log_execution(f"No Excel export link for revision {revision_id}, using CSV fallback (first sheet only)", status='warning')
                        # For now, skip CSV fallback and raise error
                        # Can be implemented later if needed
                        raise ValueError(f"No Excel export link available for revision {revision_id} (CSV fallback not implemented)")
                    else:
                        raise ValueError(f"No export links available for revision {revision_id}")

                # Download the Excel content using the credentials
                response = self.drive_service._http.request(xlsx_url)

                # Get Excel bytes
                excel_bytes = response[1]

                # Check if response is HTML (error page) vs actual content
                if isinstance(excel_bytes, bytes):
                    # Check for HTML error pages
                    try:
                        content_str = excel_bytes.decode('utf-8', errors='ignore')[:500]
                        if content_str.startswith('<!DOCTYPE') or content_str.startswith('<html'):
                            # This is an error page, likely rate limiting
                            if 'Too Many Requests' in content_str or '429' in content_str:
                                error_msg = "Rate limited (429 Too Many Requests)"
                            elif 'unavailable' in content_str.lower():
                                error_msg = "Service unavailable"
                            else:
                                error_msg = f"HTML error response (not Excel content)"

                            if attempt < max_retries - 1:
                                # Exponential backoff: 1s, 8s, 16s
                                wait_time = 4 ** attempt
                                log_execution(f"{error_msg}, retrying in {wait_time}s (attempt {attempt + 1}/{max_retries})", status='working')
                                time.sleep(wait_time)
                                continue
                            else:
                                raise ValueError(f"Failed after {max_retries} retries: {error_msg}")
                    except UnicodeDecodeError:
                        # Binary content, not HTML - proceed with Excel parsing
                        pass

                # Parse Excel with openpyxl
                workbook = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)

                # Extract data from all sheets
                sheets_data = []
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]

                    # Convert worksheet to 2D array of values, filtering empty rows
                    values = []
                    for row in worksheet.iter_rows(values_only=True):
                        # Convert None to empty string and all values to strings
                        row_values = [str(cell) if cell is not None else '' for cell in row]

                        # Skip completely empty rows
                        if not is_empty_row(row_values):
                            values.append(row_values)

                    # Trim trailing empty columns from all rows
                    if values:
                        values = trim_empty_columns(values)

                    sheets_data.append({
                        'title': sheet_name,
                        'values': values
                    })

                workbook.close()

                # Convert to markdown using shared utilities
                markdown_text = sheets_to_markdown("Revision Export", sheets_data)

                # Success - return the markdown content
                return markdown_text

            except Exception as e:
                # Check if this is a 404 error (revision doesn't exist / not exportable)
                # Google Drive API lists all revisions, but some are pruned and not exportable
                error_str = str(e)
                is_404 = ('404' in error_str and 'Revision not found' in error_str) or \
                         (hasattr(e, 'resp') and hasattr(e.resp, 'status') and e.resp.status == 404)

                if is_404:
                    # Revision isn't exportable - raise FileNotFoundError immediately without retry
                    # The caller (save_missing_revisions) will catch this and skip the revision
                    raise FileNotFoundError(f"Revision {revision_id} not available for export (pruned by Google)")

                last_error = e
                if attempt < max_retries - 1:
                    # Exponential backoff for general errors too
                    wait_time = 4 ** attempt
                    log_execution(f"Error exporting revision {revision_id}: {e}, retrying in {wait_time}s (attempt {attempt + 1}/{max_retries})", status='working')
                    time.sleep(wait_time)
                else:
                    log_execution(f"Error exporting revision {revision_id}: {e}", status='error')

        # All retries exhausted
        raise last_error if last_error else Exception(f"Failed to export revision {revision_id}")

    def _load_saved_revision_ids_from_metadata(self, metadata_path: Path) -> set:
        """
        Load all Google revision IDs that have been previously saved from metadata.jsonl.

        Reads all entries in the metadata.jsonl file and extracts revision_ids from
        their edit_history to determine which revisions are already exported.

        Args:
            metadata_path: Path to the .metadata.jsonl file

        Returns:
            Set of Google revision_id strings that have been previously saved
        """
        saved_revision_ids = set()

        if not metadata_path.exists():
            return saved_revision_ids

        try:
            with open(metadata_path, 'r', encoding='utf-8') as f:
                for line in f:
                    if not line.strip():
                        continue
                    try:
                        entry = json.loads(line)
                        # Extract revision_ids from edit_history
                        edit_history = entry.get('edit_history', [])
                        for edit in edit_history:
                            revision_id = edit.get('revision_id')
                            if revision_id:
                                saved_revision_ids.add(revision_id)

                        # Also check for revisions_exported from previous saves
                        # This handles the case where revisions were already exported
                        revisions_exported = entry.get('revisions_exported', {})
                        revision_mappings = revisions_exported.get('revision_mappings', [])
                        for mapping in revision_mappings:
                            revision_id = mapping.get('google_revision_id')
                            if revision_id:
                                saved_revision_ids.add(revision_id)
                    except json.JSONDecodeError:
                        log_execution(f"Could not parse metadata line: {line[:100]}", status='error')
                        continue

        except Exception as e:
            log_execution(f"Could not read metadata file: {e}", status='error')

        return saved_revision_ids

    def _get_next_sequence_number(self, metadata_path: Path) -> int:
        """
        Get the next sequence number to use for revision files.

        Reads revision_mappings from metadata.jsonl (source of truth) and returns
        the highest sequence number + 1. If metadata doesn't exist, starts from 1.

        Args:
            metadata_path: Path to the .metadata.jsonl file

        Returns:
            Next sequence number to use (1-based)
        """
        if not metadata_path.exists():
            return 1

        max_seq = 0
        try:
            with open(metadata_path, 'r', encoding='utf-8') as f:
                for line in f:
                    if not line.strip():
                        continue
                    try:
                        entry = json.loads(line)
                        # Extract sequence numbers from revisions_exported
                        revisions_exported = entry.get('revisions_exported', {})
                        revision_mappings = revisions_exported.get('revision_mappings', [])
                        for mapping in revision_mappings:
                            seq_str = mapping.get('sequence')
                            if seq_str:
                                try:
                                    seq_num = int(seq_str)
                                    max_seq = max(max_seq, seq_num)
                                except ValueError:
                                    pass
                    except json.JSONDecodeError:
                        continue
        except Exception:
            pass

        return max_seq + 1

    def save_missing_revisions(
        self,
        document_name: str,
        spreadsheet_id: str,
        output_path: str,
        revisions_metadata: List[Dict[str, Any]],
        max_count: int = 100,
        metadata_path: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Save only missing revisions (ones not already in the folder).

        Uses metadata.jsonl as the source of truth for tracking which Google
        revision_ids have already been exported. On restart, reads the metadata
        file to rebuild the saved set and continue sequence numbering from where
        it left off.

        Args:
            document_name: Document name for folder naming
            spreadsheet_id: Google Sheets spreadsheet ID
            output_path: Path where to save revisions
            revisions_metadata: List of revision metadata from get_edit_history
            max_count: Maximum revisions to save
            metadata_path: Path to the .metadata.jsonl file for state persistence

        Returns:
            Dict with:
            - files_created: List of newly created file paths
            - revisions_exported: Dict with first/last sequence and revision mappings (for metadata)
        """
        output_path = Path(output_path)
        revisions_dir = output_path / f"{document_name}.revisions"
        revisions_dir.mkdir(parents=True, exist_ok=True)

        # Load saved revision IDs from metadata.jsonl (source of truth)
        metadata_file = Path(metadata_path) if metadata_path else output_path / f"{document_name}.metadata.jsonl"
        saved_revision_ids = self._load_saved_revision_ids_from_metadata(metadata_file)
        log_execution(f"Previously saved revisions from metadata: {len(saved_revision_ids)}", status='info')

        # Get next sequence number from metadata.jsonl (source of truth)
        # If metadata doesn't exist, starts from 1 (ensures consistent numbering on fresh start)
        next_seq = self._get_next_sequence_number(metadata_file)

        files_created = []
        revision_mappings = []
        first_sequence = None
        import time

        for i, rev_metadata in enumerate(revisions_metadata):
            # Handle both 'revision_id' (from get_edit_history) and 'id' keys
            revision_id = rev_metadata.get('revision_id') or rev_metadata.get('id')

            # Skip if this revision_id has already been saved
            if revision_id in saved_revision_ids:
                log_execution(f"Revision {revision_id} already saved", status='info')
                continue

            # Format: {timestamp}-{seq}.md
            timestamp = rev_metadata.get('timestamp', '')
            # Parse ISO timestamp and extract date/time
            if timestamp:
                # Convert ISO 8601 to simple format: 2025-10-28-143200
                dt = datetime.fromisoformat(timestamp.replace('+00:00', ''))
                timestamp_str = dt.strftime('%Y-%m-%d-%H%M%S')
            else:
                timestamp_str = 'unknown'

            seq_str = f"{next_seq:03d}"
            rev_file = revisions_dir / f"{timestamp_str}-{seq_str}.md"

            try:
                content = self.export_revision(spreadsheet_id, revision_id)
                rev_file.write_text(content, encoding='utf-8')
                files_created.append(str(rev_file))

                # Track this revision mapping for metadata
                file_size = len(content.encode())
                content_hash = hashlib.sha256(content.encode()).hexdigest()

                mapping = {
                    'sequence': seq_str,
                    'google_revision_id': revision_id,
                    'exported_filename': rev_file.name,
                    'file_size_bytes': file_size,
                    'content_hash': f"sha256:{content_hash}"
                }
                revision_mappings.append(mapping)

                if first_sequence is None:
                    first_sequence = seq_str
                last_sequence = seq_str

                saved_revision_ids.add(revision_id)
                next_seq += 1
                log_execution(f"Saved revision {seq_str}: {rev_file.name}", status='success')

                # Add delay between requests to avoid rate limiting (except for last revision)
                if i < len(revisions_metadata) - 1:
                    time.sleep(2.0)  # 2 second delay between revisions to avoid 429 errors

            except FileNotFoundError as e:
                # Revision not exportable (pruned by Google) - skip it and continue
                log_execution(f"Skipping revision {revision_id}: {e}", status='warning')
                # Mark this as saved so we don't try again
                saved_revision_ids.add(revision_id)
                # Don't increment next_seq - we didn't actually save a file
                continue

            except Exception as e:
                # Other errors - log and continue
                log_execution(f"Failed to save revision {seq_str}: {e}", status='error')
                # Don't mark as saved - we might retry it later
                continue

        # Build the revisions_exported structure for metadata
        revisions_exported = {}
        if revision_mappings:
            revisions_exported = {
                'first_sequence': first_sequence,
                'last_sequence': last_sequence,
                'revision_mappings': revision_mappings
            }
            log_execution(f"Exported {len(revision_mappings)} new revisions (seq {first_sequence}-{last_sequence})", status='success')
        else:
            log_execution(f"No new revisions to export", status='info')

        return {
            'files_created': files_created,
            'revisions_exported': revisions_exported
        }
