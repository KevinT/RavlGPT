#!/usr/bin/env python3
"""
Google Sheets Analyzer Workflow

Handles extraction and analysis of data from Google Sheets and Excel files.

Inherits from GoogleAPIsMixin to access Google Sheets service.
"""

import re
import sys
from datetime import datetime
from typing import Dict, Any, List
from io import BytesIO


class GoogleSheetsAnalyzer:
    """
    Analyzes and extracts data from Google Sheets.

    Usage:
        analyzer = GoogleSheetsAnalyzer(loop)  # loop must inherit GoogleAPIsMixin
        sheet_data = analyzer.fetch(url)
        markdown = analyzer.export_as_markdown(url)
    """

    def __init__(self, loop_with_mixin):
        """
        Initialize with a loop that has GoogleAPIsMixin.

        Args:
            loop_with_mixin: A RAVL loop instance that inherits GoogleAPIsMixin
        """
        self.loop = loop_with_mixin

    def fetch(self, url: str) -> Dict[str, Any]:
        """
        Fetch content from Google Sheets using the API.

        Args:
            url: Google Sheets URL

        Returns:
            Dict with:
            - 'text': Markdown formatted representation of all sheets
            - 'title': Spreadsheet title
            - 'last_modified': ISO timestamp of last modification
            - 'spreadsheet_id': Extracted spreadsheet ID from URL
            - 'sheets': List of sheet data (name, values)

        Raises:
            ValueError: If URL is invalid
            Exception: If API request fails
        """
        if not self.loop.google_sheets_service or not self.loop.google_drive_service:
            self.loop.init_google_services()

        # Extract spreadsheet ID from URL
        # Format: https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit...
        spreadsheet_id_match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
        if not spreadsheet_id_match:
            raise ValueError(f"Invalid Google Sheets URL: {url}")

        spreadsheet_id = spreadsheet_id_match.group(1)

        try:
            # Detect file type using Drive API
            file_metadata = self.loop.google_drive_service.files().get(
                fileId=spreadsheet_id,
                fields='mimeType,name,modifiedTime',
                supportsAllDrives=True
            ).execute()

            mime_type = file_metadata.get('mimeType')

            # Branch based on file type
            if mime_type == 'application/vnd.google-apps.spreadsheet':
                # Native Google Sheets - use Sheets API
                return self._fetch_native_sheets(spreadsheet_id, file_metadata)
            elif mime_type in [
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
                'application/vnd.ms-excel'  # .xls
            ]:
                # Excel file - download and parse with openpyxl
                return self._fetch_excel_file(spreadsheet_id, file_metadata)
            else:
                raise ValueError(f"Unsupported spreadsheet MIME type: {mime_type}")

        except Exception as e:
            raise Exception(f"Failed to fetch spreadsheet: {e}")

    def _fetch_native_sheets(self, spreadsheet_id: str, file_metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Fetch native Google Sheets using Sheets API.

        Args:
            spreadsheet_id: The spreadsheet ID
            file_metadata: File metadata from Drive API

        Returns:
            Dict with text, title, last_modified, spreadsheet_id, sheets
        """
        try:
            # Get spreadsheet metadata and all sheets
            spreadsheet = self.loop.google_sheets_service.spreadsheets().get(
                spreadsheetId=spreadsheet_id
            ).execute()

            title = spreadsheet.get('properties', {}).get('title', 'Untitled Spreadsheet')
            sheets = spreadsheet.get('sheets', [])

            # Get modification timestamp from Drive API metadata
            last_modified = file_metadata.get('modifiedTime')

            # Fetch data for each sheet
            sheets_data = []
            for sheet in sheets:
                sheet_properties = sheet.get('properties', {})
                sheet_title = sheet_properties.get('title', 'Untitled Sheet')
                sheet_id = sheet_properties.get('sheetId')

                try:
                    # Get all values for this sheet
                    result = self.loop.google_sheets_service.spreadsheets().values().get(
                        spreadsheetId=spreadsheet_id,
                        range=sheet_title  # Use sheet title as range to get all data
                    ).execute()

                    values = result.get('values', [])

                    sheets_data.append({
                        'title': sheet_title,
                        'sheet_id': sheet_id,
                        'values': values
                    })

                except Exception as e:
                    from pathlib import Path
                    _utils_dir = Path(__file__).parent.parent / 'utils'
                    import sys
                    if str(_utils_dir) not in sys.path:
                        sys.path.insert(0, str(_utils_dir))
                    from logging_utils import log_execution
                    log_execution(f"Could not fetch data for sheet '{sheet_title}': {e}", status='error')
                    sheets_data.append({
                        'title': sheet_title,
                        'sheet_id': sheet_id,
                        'values': [],
                        'error': str(e)
                    })

            # Convert to markdown
            markdown_text = self._sheets_to_markdown(title, sheets_data)

            return {
                'text': markdown_text,
                'title': title,
                'last_modified': last_modified,
                'spreadsheet_id': spreadsheet_id,
                'sheets': sheets_data
            }

        except Exception as e:
            raise Exception(f"Failed to fetch native Google Sheets: {e}")

    def _fetch_excel_file(self, spreadsheet_id: str, file_metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Fetch Excel file (.xlsx/.xls) from Google Drive and parse with openpyxl.

        Args:
            spreadsheet_id: The file ID in Google Drive
            file_metadata: File metadata from Drive API

        Returns:
            Dict with text, title, last_modified, spreadsheet_id, sheets
        """
        try:
            # Import openpyxl
            try:
                import openpyxl
            except ImportError:
                raise Exception(
                    "openpyxl is required to parse Excel files. "
                    "Add it to allowed_dependencies in config/ravl.yml"
                )

            # Download Excel file as binary
            request = self.loop.google_drive_service.files().get_media(
                fileId=spreadsheet_id,
                supportsAllDrives=True
            )
            excel_bytes = request.execute()

            # Parse with openpyxl
            workbook = openpyxl.load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)

            title = file_metadata.get('name', 'Untitled Spreadsheet')
            last_modified = file_metadata.get('modifiedTime')

            # Extract data from all sheets
            sheets_data = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Convert worksheet to 2D array of values, filtering empty rows
                values = []
                for row in worksheet.iter_rows(values_only=True):
                    # Convert None to empty string and all values to strings
                    row_values = [str(cell) if cell is not None else '' for cell in row]

                    # Skip completely empty rows
                    if not self._is_empty_row(row_values):
                        values.append(row_values)

                # Trim trailing empty columns from all rows
                if values:
                    values = self._trim_empty_columns(values)

                sheets_data.append({
                    'title': sheet_name,
                    'sheet_id': None,  # Excel doesn't have sheet IDs
                    'values': values
                })

            workbook.close()

            # Convert to markdown
            markdown_text = self._sheets_to_markdown(title, sheets_data)

            return {
                'text': markdown_text,
                'title': title,
                'last_modified': last_modified,
                'spreadsheet_id': spreadsheet_id,
                'sheets': sheets_data
            }

        except Exception as e:
            raise Exception(f"Failed to fetch Excel file: {e}")

    def _is_empty_row(self, row: List[str]) -> bool:
        """
        Check if a row contains only empty or whitespace strings.

        Args:
            row: List of cell values as strings

        Returns:
            True if row is completely empty, False otherwise
        """
        return all(cell.strip() == '' for cell in row)

    def _trim_empty_columns(self, values: List[List[str]]) -> List[List[str]]:
        """
        Remove trailing empty columns from all rows in a dataset.

        Args:
            values: 2D list of cell values

        Returns:
            2D list with trailing empty columns removed
        """
        if not values:
            return values

        # Find the maximum column index that contains non-empty data
        max_col_with_data = 0
        for row in values:
            for i, cell in enumerate(row):
                if cell.strip() != '':
                    max_col_with_data = max(max_col_with_data, i)

        # Trim all rows to max_col_with_data + 1 (convert index to count)
        trimmed_values = [row[:max_col_with_data + 1] for row in values]

        return trimmed_values

    def export_as_markdown(self, url: str) -> str:
        """
        Fetch sheets and return as markdown string.

        Args:
            url: Google Sheets URL

        Returns:
            Markdown formatted string

        Raises:
            ValueError: If URL is invalid
            Exception: If API request fails
        """
        sheet_data = self.fetch(url)
        return sheet_data['text']

    def _sheets_to_markdown(self, title: str, sheets_data: List[Dict[str, Any]]) -> str:
        """
        Convert sheets data to markdown format.

        Args:
            title: Spreadsheet title
            sheets_data: List of sheet data dicts

        Returns:
            Markdown formatted string
        """
        markdown_parts = [f"# {title}\n"]

        for sheet in sheets_data:
            sheet_title = sheet.get('title', 'Untitled Sheet')
            values = sheet.get('values', [])
            error = sheet.get('error')

            markdown_parts.append(f"\n## {sheet_title}\n")

            if error:
                markdown_parts.append(f"*Error fetching sheet data: {error}*\n")
                continue

            if not values:
                markdown_parts.append("*No data in this sheet*\n")
                continue

            # Convert to markdown table
            markdown_parts.append(self._values_to_markdown_table(values))

        return "\n".join(markdown_parts)

    def _values_to_markdown_table(self, values: List[List[str]]) -> str:
        """
        Convert 2D array of values to markdown table.

        Args:
            values: 2D list of cell values

        Returns:
            Markdown table string
        """
        if not values:
            return ""

        # Find maximum number of columns
        max_cols = max(len(row) for row in values) if values else 0

        if max_cols == 0:
            return ""

        # Normalize all rows to have same number of columns (pad with empty strings)
        normalized_rows = []
        for row in values:
            normalized_row = list(row) + [''] * (max_cols - len(row))
            normalized_rows.append(normalized_row)

        # Calculate column widths
        col_widths = [0] * max_cols
        for row in normalized_rows:
            for i, cell in enumerate(row):
                col_widths[i] = max(col_widths[i], len(str(cell)))

        # Minimum width of 3 for each column
        col_widths = [max(3, width) for width in col_widths]

        # Build markdown table
        lines = []

        # Header row (first row of data)
        if normalized_rows:
            header_cells = [str(cell).ljust(col_widths[i]) for i, cell in enumerate(normalized_rows[0])]
            lines.append("| " + " | ".join(header_cells) + " |")

            # Separator row
            separators = ["-" * col_widths[i] for i in range(max_cols)]
            lines.append("| " + " | ".join(separators) + " |")

            # Data rows
            for row in normalized_rows[1:]:
                row_cells = [str(cell).ljust(col_widths[i]) for i, cell in enumerate(row)]
                lines.append("| " + " | ".join(row_cells) + " |")

        return "\n".join(lines) + "\n"
